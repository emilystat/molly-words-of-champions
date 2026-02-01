#!/usr/bin/env python3
"""
Verify words from Excel against the official PDF and identify discrepancies.
"""

import re
from openpyxl import load_workbook

try:
    import PyPDF2
    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False
    print("PyPDF2 not available, will try pdfplumber")

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False
    print("pdfplumber not available")

def extract_words_from_pdf(pdf_path):
    """Extract words from the PDF file."""
    words = {
        'one_bee': [],
        'two_bee': [],
        'three_bee': []
    }

    if HAS_PDFPLUMBER:
        with pdfplumber.open(pdf_path) as pdf:
            current_section = None

            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue

                # Check for section headers
                if 'one bee' in text.lower() or 'one-bee' in text.lower():
                    current_section = 'one_bee'
                    print(f"Found One Bee section on page {page.page_number}")
                elif 'two bee' in text.lower() or 'two-bee' in text.lower():
                    current_section = 'two_bee'
                    print(f"Found Two Bee section on page {page.page_number}")
                elif 'three bee' in text.lower() or 'three-bee' in text.lower():
                    current_section = 'three_bee'
                    print(f"Found Three Bee section on page {page.page_number}")

                # Extract word entries (word followed by definition/pronunciation)
                # Looking for patterns like: "word [pronunciation] definition"
                lines = text.split('\n')
                for line in lines:
                    line = line.strip()
                    # Skip empty lines, headers, page numbers
                    if not line or line.isdigit() or len(line) < 3:
                        continue

                    # Skip common headers
                    if any(h in line.lower() for h in ['bee', 'page', 'words of', 'champion']):
                        if line.lower().strip() not in ['bee', 'one bee', 'two bee', 'three bee']:
                            continue

                    # Try to extract the word (usually at the start of the line)
                    # Word is typically followed by pronunciation in brackets or definition
                    match = re.match(r'^([a-zA-Z\s\-\']+?)[\s]*[\[\(]', line)
                    if match and current_section:
                        word = match.group(1).strip()
                        if len(word) > 1:  # Ignore single letters
                            words[current_section].append(word)

    return words

def read_excel_words(excel_path):
    """Read words from Excel file."""
    wb = load_workbook(excel_path)

    words = {
        'one_bee': [],
        'two_bee': [],
        'three_bee': []
    }

    # Read One Bee
    if 'One Bee' in wb.sheetnames:
        ws = wb['One Bee']
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0]:  # If there's a word in column A
                word = str(row[0]).strip()
                words['one_bee'].append((row_idx, word))

    # Read Two Bee
    if 'Two Bee' in wb.sheetnames:
        ws = wb['Two Bee']
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0]:
                word = str(row[0]).strip()
                words['two_bee'].append((row_idx, word))

    # Read Three Bee
    if 'Three Bee' in wb.sheetnames:
        ws = wb['Three Bee']
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0]:
                word = str(row[0]).strip()
                words['three_bee'].append((row_idx, word))

    return words

def find_suspicious_words(excel_words):
    """Find words that look suspicious (not typical spelling bee words)."""
    suspicious = []

    for section, word_list in excel_words.items():
        for row_idx, word in word_list:
            # Check for suspicious patterns
            suspicious_patterns = [
                (r'\b(it|is|was|were|be|been|are|am)\b', 'Contains common verb'),
                (r'\b(the|a|an)\b', 'Contains article'),
                (r'\b(not|until|turned)\b', 'Contains suspicious common word'),
                (r'^[A-Z][a-z]+\s+[a-z]+', 'Starts with capital followed by lowercase word'),
                (r'[\.!?]', 'Contains punctuation'),
            ]

            for pattern, reason in suspicious_patterns:
                if re.search(pattern, word, re.IGNORECASE):
                    suspicious.append({
                        'section': section,
                        'row': row_idx,
                        'word': word,
                        'reason': reason
                    })
                    break

    return suspicious

def main():
    excel_path = '../WordsOfChampions_OneTwoThreeBee_withRoots.xlsx'
    pdf_path = '../WordsOfChampions.pdf'

    print("Reading Excel file...")
    excel_words = read_excel_words(excel_path)

    print("\nExcel word counts:")
    for section, words in excel_words.items():
        print(f"  {section}: {len(words)} words")

    print("\nFinding suspicious words...")
    suspicious = find_suspicious_words(excel_words)

    if suspicious:
        print(f"\nFound {len(suspicious)} suspicious entries:")
        for item in suspicious:
            print(f"  [{item['section']}] Row {item['row']}: '{item['word']}' - {item['reason']}")
    else:
        print("\nNo suspicious words found!")

    # Try to extract from PDF if available
    if HAS_PDFPLUMBER:
        print("\n" + "="*60)
        print("Extracting words from PDF...")
        pdf_words = extract_words_from_pdf(pdf_path)
        print("\nPDF word counts:")
        for section, words in pdf_words.items():
            print(f"  {section}: {len(words)} words")
            if words:
                print(f"    First few: {words[:5]}")

if __name__ == '__main__':
    main()
