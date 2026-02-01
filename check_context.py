#!/usr/bin/env python3
"""
Check the context around the suspicious rows in the Excel file.
"""

from openpyxl import load_workbook

def check_context(excel_path, sheet_name, row_num, context_lines=5):
    """Check rows around a specific row number."""
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    print(f"\nContext around row {row_num} in {sheet_name}:")
    print("=" * 80)

    start_row = max(2, row_num - context_lines)
    end_row = row_num + context_lines + 1

    for row_idx in range(start_row, end_row):
        row = list(ws[row_idx])
        word = row[0].value if row[0].value else ""
        marker = " <--- SUSPICIOUS" if row_idx == row_num else ""
        print(f"Row {row_idx}: {word}{marker}")

def main():
    excel_path = '../WordsOfChampions_OneTwoThreeBee_withRoots.xlsx'

    # Check context for the two suspicious entries
    check_context(excel_path, 'One Bee', 572, context_lines=5)
    check_context(excel_path, 'One Bee', 869, context_lines=5)

if __name__ == '__main__':
    main()
