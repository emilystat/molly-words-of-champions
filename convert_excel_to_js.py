#!/usr/bin/env python3
"""
Convert WordsOfChampions Excel file to JavaScript data files for the website
"""

import openpyxl
import json

# Load the Excel file
excel_path = '../WordsOfChampions_OneTwoThreeBee_withRoots.xlsx'
wb = openpyxl.load_workbook(excel_path, read_only=True)

# Function to convert words from a sheet to JavaScript format
def convert_words_sheet(sheet_name, word_list_level, bee_difficulty):
    ws = wb[sheet_name]
    words = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        word = row[0]
        if word:  # Skip empty rows
            word_obj = {
                "word": word.strip(),
                "definition": "[Definition to be added]",
                "sentence": f"[Example sentence for '{word}' to be added]",
                "wordList": word_list_level,
                "beeDifficulty": bee_difficulty,
                "languageOrigin": "To be determined",
                "roots": [],
                "isNewThisYear": False,
                "pronunciationGuide": "",
                "audioUrl": None
            }
            words.append(word_obj)

    return words

# Function to convert roots to JavaScript format
def convert_roots_sheet(sheet_name, language_family):
    ws = wb[sheet_name]
    roots = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # If root exists
            root_obj = {
                "root": row[0].strip() if row[0] else "",
                "meaning": row[1].strip() if row[1] else "",
                "languageFamily": language_family,
                "examples": row[2].strip() if row[2] else ""
            }
            roots.append(root_obj)

    return roots

# Convert word sheets
print("Converting word sheets...")
words_1000 = convert_words_sheet('One Bee', '1000', 'oneBee')
words_2000 = convert_words_sheet('Two Bee', '2000', 'twoBee')
words_3000 = convert_words_sheet('Three Bee', '3000', 'threeBee')

print(f"  One Bee: {len(words_1000)} words")
print(f"  Two Bee: {len(words_2000)} words")
print(f"  Three Bee: {len(words_3000)} words")

# Convert roots sheets
print("\nConverting roots sheets...")
greek_roots = convert_roots_sheet('Greek Roots', 'Greek')
latin_roots = convert_roots_sheet('Latin Roots', 'Latin')

print(f"  Greek Roots: {len(greek_roots)} roots")
print(f"  Latin Roots: {len(latin_roots)} roots")

# Function to write JavaScript file
def write_js_file(filename, var_name, data, comment):
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(f"// {comment}\n")
        f.write(f"// Auto-generated from Excel file\n")
        f.write(f"// Total items: {len(data)}\n\n")
        f.write(f"const {var_name} = ")
        f.write(json.dumps(data, indent=2, ensure_ascii=False))
        f.write(";\n")

# Write words files
print("\nWriting JavaScript data files...")
write_js_file(
    'data/words-1000.js',
    'words1000',
    words_1000,
    'One Bee: 1,000 words (easier words with common patterns)'
)

write_js_file(
    'data/words-2000.js',
    'words2000',
    words_2000,
    'Two Bee: 2,000 words (trickier words with more complex patterns)'
)

write_js_file(
    'data/words-3000.js',
    'words3000',
    words_3000,
    'Three Bee: 1,000 words (challenging words with irregular patterns)'
)

# Combine all roots
all_roots = greek_roots + latin_roots

write_js_file(
    'data/roots-data.js',
    'rootsData',
    all_roots,
    'Roots and Patterns Library - Greek and Latin roots'
)

wb.close()

print("\n✅ Conversion complete!")
print(f"\nTotal words: {len(words_1000) + len(words_2000) + len(words_3000)}")
print(f"Total roots: {len(all_roots)}")
print("\nGenerated files:")
print("  - data/words-1000.js")
print("  - data/words-2000.js")
print("  - data/words-3000.js")
print("  - data/roots-data.js")
