#!/usr/bin/env python3
"""
Update all word lists (One Bee, Two Bee, Three Bee) from Excel files
"""

import openpyxl
import json

def create_word_object(word, word_list, bee_difficulty):
    """Create a word object with Google links for missing definitions"""
    return {
        "word": word.strip(),
        "wordList": word_list,
        "beeDifficulty": bee_difficulty,
        "languageOrigin": "To be determined",
        "roots": [],
        "isNewThisYear": False,
        "pronunciationGuide": "",
        "audioUrl": None,
        "definition": f'[Definition not available. <a href="https://www.google.com/search?q={word.strip()}+meaning" target="_blank">Search Google</a>]',
        "googleLink": f"https://www.google.com/search?q={word.strip()}+meaning",
        "sentence": f'[Example sentence to be added. <a href="https://www.google.com/search?q={word.strip()}+example+sentence" target="_blank">Search Google</a>]'
    }

def extract_words_from_sheet(wb, sheet_name, word_list, bee_difficulty):
    """Extract words from a sheet in the workbook"""
    ws = wb[sheet_name]
    words = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # If word exists
            word = row[0]
            words.append(create_word_object(word, word_list, bee_difficulty))

    return words

def write_js_file(filename, var_name, data, comment):
    """Write JavaScript file with word data"""
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(f"// {comment}\n")
        f.write(f"// Auto-generated from Excel files\n")
        f.write(f"// Total items: {len(data)}\n")
        f.write(f"// With definitions: 0, Google links: {len(data)}\n")
        f.write("\n")
        f.write(f"const {var_name} = ")
        f.write(json.dumps(data, indent=2, ensure_ascii=False))
        f.write(";\n")

print("📚 Updating all word lists from Excel files...\n")

# Update One Bee and Three Bee from WordsOfChampions_OneTwoThreeBee_withRoots.xlsx
print("Reading WordsOfChampions_OneTwoThreeBee_withRoots.xlsx...")
wb1 = openpyxl.load_workbook('../WordsOfChampions_OneTwoThreeBee_withRoots.xlsx', read_only=True)

print("  Extracting One Bee words...")
words_1000 = extract_words_from_sheet(wb1, 'One Bee', '1000', 'oneBee')
print(f"  ✓ One Bee: {len(words_1000)} words")

print("  Extracting Three Bee words...")
words_3000 = extract_words_from_sheet(wb1, 'Three Bee', '3000', 'threeBee')
print(f"  ✓ Three Bee: {len(words_3000)} words")

wb1.close()

# Update Two Bee from TwoBee_AllWords_Alphabetical.xlsx
print("\nReading TwoBee_AllWords_Alphabetical.xlsx...")
wb2 = openpyxl.load_workbook('../TwoBee_AllWords_Alphabetical.xlsx', read_only=True)

print("  Extracting Two Bee words...")
words_2000 = extract_words_from_sheet(wb2, 'Two Bee', '2000', 'twoBee')
print(f"  ✓ Two Bee: {len(words_2000)} words")

wb2.close()

# Write JavaScript files
print("\n📝 Writing JavaScript data files...")

write_js_file(
    'data/words-1000.js',
    'words1000',
    words_1000,
    'One Bee: 1,000 words (easier words with common patterns)'
)
print("  ✓ data/words-1000.js")

write_js_file(
    'data/words-2000.js',
    'words2000',
    words_2000,
    'Two Bee: 2,000 words (trickier words with more complex patterns)'
)
print("  ✓ data/words-2000.js")

write_js_file(
    'data/words-3000.js',
    'words3000',
    words_3000,
    'Three Bee: 1,000 words (challenging words with irregular patterns)'
)
print("  ✓ data/words-3000.js")

print("\n✅ All word lists updated successfully!")
print(f"\nTotal words: {len(words_1000) + len(words_2000) + len(words_3000)}")
print(f"  - One Bee: {len(words_1000)} words")
print(f"  - Two Bee: {len(words_2000)} words")
print(f"  - Three Bee: {len(words_3000)} words")
